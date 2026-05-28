"""
generate_excel.py  —  Job Hunting Machine (OpenRouter edition)
--------------------------------------------------------------
1. Reads resume.pdf or resume.txt from repo root
2. Reads prompt.txt, substitutes placeholders
3. Calls OpenRouter free-tier LLMs with retry + backoff
4. Parses JSON response robustly (handles control chars, markdown fences)
5. Writes a rich, deduplicated Excel workbook:
      Sheet 1 – Dashboard (summary + strengths/gaps)
      Sheet 2 – All Jobs (sorted by fit_score, deduplicated)
      Sheet 3 – High Probability  (fit >= 75)
      Sheet 4 – Medium Probability (50–74)
      Sheet 5 – Stretch Roles      (< 50)
      Sheet 6 – 30-Day Resume Tips

Required GitHub Secret:
  OPENROUTER_API_KEY  — free key from https://openrouter.ai
"""

import os, json, re, time, datetime, glob, sys
import urllib.request, urllib.error
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── OpenRouter model list (tried in order, all free tier) ────────────────────
# If one is rate-limited or errors, the next is tried automatically.
FREE_MODELS = [
    "openai/gpt-oss-120b:free",
    "deepseek/deepseek-v4-flash:free",
    "meta-llama/llama-3.3-70b-instruct:free",
    "google/gemma-4-31b-it:free",
    "mistralai/mistral-7b-instruct:free",
    "microsoft/phi-4-reasoning-plus:free",
]

OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"


# ── styling helpers ───────────────────────────────────────────────────────────

def hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))

def bfont(size=11, color="FFFFFF", bold=True, italic=False):
    return Font(bold=bold, italic=italic, size=size, color=color.lstrip("#"))

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

PROB_COLORS = {
    "High":    ("1B5E20", "C8E6C9"),
    "Medium":  ("E65100", "FFE0B2"),
    "Stretch": ("B71C1C", "FFCDD2"),
}

SCORE_FILL = {
    (75, 101): "C8E6C9",
    (50,  75): "FFE0B2",
    (0,   50): "FFCDD2",
}

def score_fill(score):
    for (lo, hi), color in SCORE_FILL.items():
        if lo <= score < hi:
            return hfill(color)
    return hfill("FFFFFF")

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def header_row(ws, headers, fill_hex, font_color="FFFFFF"):
    ws.append(headers)
    row = ws.max_row
    for col in range(1, len(headers)+1):
        c = ws.cell(row=row, column=col)
        c.fill      = hfill(fill_hex)
        c.font      = bfont(10, font_color)
        c.alignment = center(wrap=True)
        c.border    = thin_border()
    ws.row_dimensions[row].height = 22


# ── resume reader ─────────────────────────────────────────────────────────────

def read_resume():
    """Return resume text extracted from PDF or plain text file."""
    for pattern in ["resume.pdf", "resume.PDF", "resume.txt",
                    "cv.pdf",     "cv.PDF",     "cv.txt"]:
        matches = glob.glob(pattern)
        if matches:
            path = matches[0]
            if path.lower().endswith(".pdf"):
                return extract_pdf_text(path)
            else:
                with open(path, "r", encoding="utf-8", errors="replace") as f:
                    text = f.read()
                print(f"📄 Resume loaded ({len(text)} chars)")
                return text
    raise FileNotFoundError(
        "No resume found! Add resume.pdf or resume.txt to the repo root."
    )


def extract_pdf_text(path):
    """Extract text from PDF using pdfminer.six (installed in workflow)."""
    try:
        from pdfminer.high_level import extract_text
        text = extract_text(path)
        text = text.strip()
        print(f"✅ PDF extracted ({len(text)} chars)")
        return text
    except ImportError:
        # Fallback: read raw bytes and decode printable ASCII
        print("⚠️  pdfminer not available, using fallback PDF reader")
        with open(path, "rb") as f:
            raw = f.read()
        text = raw.decode("latin-1", errors="replace")
        # Keep only printable lines
        lines = [l for l in text.splitlines()
                 if any(c.isalpha() for c in l)]
        result = "\n".join(lines[:300])
        print(f"✅ PDF fallback extracted ({len(result)} chars)")
        return result


# ── robust JSON parser ────────────────────────────────────────────────────────

def clean_and_parse(raw: str) -> dict:
    """
    Parse JSON from LLM output robustly:
    - strips markdown fences
    - removes unescaped control characters
    - extracts first {...} block if there's preamble
    """
    if raw is None:
        raise ValueError("Response content is None")

    raw = raw.strip()

    # Strip markdown fences  ```json ... ``` or ``` ... ```
    if raw.startswith("```"):
        parts = raw.split("```")
        # parts[1] is the content between first pair of fences
        raw = parts[1] if len(parts) > 1 else raw
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.strip()

    # Extract first JSON object if there's preamble text before it
    brace_start = raw.find("{")
    brace_end   = raw.rfind("}")
    if brace_start > 0 and brace_end > brace_start:
        raw = raw[brace_start:brace_end+1]

    # Remove unescaped control characters (tabs/newlines inside string values)
    # This fixes the "Invalid control character" error from some models
    raw = re.sub(r'(?<!\\)[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', ' ', raw)

    return json.loads(raw)


# ── OpenRouter API call with retry ───────────────────────────────────────────

def call_openrouter(prompt: str, call_label: str = "") -> tuple:
    """
    Try each free model in FREE_MODELS.
    For each model, retry up to 3 times with exponential backoff on 429s.
    Returns (parsed_dict, model_name_used).
    Raises RuntimeError if all models fail.
    """
    api_key = os.environ.get("OPENROUTER_API_KEY", "")
    if not api_key:
        raise EnvironmentError(
            "OPENROUTER_API_KEY secret is not set. "
            "Get a free key at https://openrouter.ai and add it as a GitHub secret."
        )

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type":  "application/json",
        "HTTP-Referer":  "https://github.com/job-hunting-pipeline",
        "X-Title":       "Job Hunting Pipeline",
    }

    for model in FREE_MODELS:
        for attempt in range(3):
            wait = 15 * (attempt + 1)   # 15s, 30s, 45s
            print(f"📡 [{call_label}] Trying {model} (attempt {attempt+1})...")

            payload = json.dumps({
                "model": model,
                "max_tokens": 8000,
                "temperature": 0.3,
                "messages": [
                    {
                        "role": "system",
                        "content": (
                            "You are a precise JSON API. "
                            "Return ONLY valid JSON — no markdown, no explanation, "
                            "no preamble. All string values must be properly escaped."
                        )
                    },
                    {"role": "user", "content": prompt}
                ]
            }).encode("utf-8")

            try:
                req  = urllib.request.Request(OPENROUTER_URL,
                                              data=payload,
                                              headers=headers,
                                              method="POST")
                with urllib.request.urlopen(req, timeout=120) as resp:
                    result = json.loads(resp.read().decode("utf-8"))

                content = (result.get("choices", [{}])[0]
                                 .get("message", {})
                                 .get("content"))

                if content is None:
                    # Upstream error object instead of a choices array
                    err = result.get("error", {})
                    print(f"⚠️  {model} upstream error: {err} — retrying in {wait}s...")
                    time.sleep(wait)
                    continue

                # Preview first 120 chars for debugging
                print(f"🔍 Raw preview: {content[:120].strip()}")

                parsed = clean_and_parse(content)
                print(f"✅ [{call_label}] Success with {model}")
                return parsed, model

            except urllib.error.HTTPError as e:
                body = e.read().decode("utf-8", errors="replace")[:300]
                if e.code == 429:
                    print(f"⚠️  {model} rate-limited (429) — waiting {wait}s... {body}")
                    time.sleep(wait)
                    continue
                else:
                    print(f"⚠️  {model} HTTP {e.code}: {body} — trying next model")
                    break   # non-429 HTTP error: skip to next model immediately

            except urllib.error.URLError as e:
                print(f"⚠️  {model} network error: {e.reason} — retrying in {wait}s...")
                time.sleep(wait)
                continue

            except json.JSONDecodeError as e:
                print(f"⚠️  {model} JSON parse error: {e} — trying next model")
                break   # bad output won't improve on retry; try next model

            except Exception as e:
                print(f"⚠️  {model} unexpected error: {e} — retrying in {wait}s...")
                time.sleep(wait)
                continue

        # All 3 attempts for this model exhausted
        print(f"❌ {model} exhausted all attempts — moving to next model")

    raise RuntimeError(
        "All OpenRouter free models failed. "
        "Try re-running the workflow — rate limits reset hourly. "
        "Or shift the cron to 2 AM IST (20:30 UTC) for less congestion."
    )


# ── Excel builder ─────────────────────────────────────────────────────────────

JOB_HEADERS = [
    "Rank", "Company", "Role", "Sector", "Size",
    "Location", "Exp. Required", "Key Skills",
    "Fit Score", "Probability", "Match Reason",
    "Gap / Stretch Reason", "Apply URL", "Source", "Est. CTC (LPA)"
]
JOB_WIDTHS = [6, 22, 28, 14, 12, 16, 13, 34, 10, 11, 36, 30, 40, 14, 14]


def write_job_row(ws, job, row_fill_hex=None):
    skills = ", ".join(job.get("key_skills_required", []))
    prob   = job.get("probability_category", "Medium")
    score  = int(job.get("fit_score", 0))

    row = [
        job.get("rank", ""),
        job.get("company", ""),
        job.get("role", ""),
        job.get("sector", ""),
        job.get("company_size", ""),
        job.get("location", ""),
        job.get("experience_required", ""),
        skills,
        score,
        prob,
        job.get("match_reason", ""),
        job.get("gap_reason", ""),
        job.get("apply_url", ""),
        job.get("source", ""),
        job.get("estimated_ctc_lpa", "Unknown"),
    ]
    ws.append(row)
    r = ws.max_row
    ws.row_dimensions[r].height = 42

    prob_text_color, prob_bg = PROB_COLORS.get(prob, ("000000", "FFFFFF"))

    for col_idx, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=col_idx)
        cell.border    = thin_border()
        cell.alignment = left() if col_idx > 2 else center()

        if col_idx == 9:        # Fit score — colour coded
            cell.fill      = score_fill(score)
            cell.font      = bfont(11, "1B1B1B")
            cell.alignment = center()
        elif col_idx == 10:     # Probability badge
            cell.fill      = hfill(prob_bg)
            cell.font      = bfont(10, prob_text_color)
            cell.alignment = center()
        elif col_idx == 13:     # Apply URL — blue underline
            cell.font = Font(color="1565C0", underline="single", size=10)
        else:
            if row_fill_hex:
                cell.fill = hfill(row_fill_hex)
            cell.font = Font(size=10, color="1B1B1B")


def build_excel(data: dict, output_path: str):
    wb = openpyxl.Workbook()

    today     = data.get("analysis_date", datetime.date.today().isoformat())
    target    = data.get("target_date", "")
    candidate = data.get("candidate_summary", "")
    strengths = data.get("resume_strengths", [])
    gaps      = data.get("resume_gaps", [])
    jobs      = data.get("jobs", [])
    tips      = data.get("daily_resume_tips", [])

    # Deduplicate jobs by (company, role) keeping highest fit_score
    seen = {}
    for j in jobs:
        key = (j.get("company", "").lower(), j.get("role", "").lower())
        if key not in seen or j.get("fit_score", 0) > seen[key].get("fit_score", 0):
            seen[key] = j
    jobs = sorted(seen.values(), key=lambda x: -x.get("fit_score", 0))
    for i, j in enumerate(jobs, 1):
        j["rank"] = i

    high    = [j for j in jobs if j.get("probability_category") == "High"]
    medium  = [j for j in jobs if j.get("probability_category") == "Medium"]
    stretch = [j for j in jobs if j.get("probability_category") == "Stretch"]

    # ── Sheet 1: Dashboard ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "📊 Dashboard"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60

    ws.append([])
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value     = f"🎯 Job Hunt Dashboard  —  {today}  →  {target}"
    c.font      = bfont(15, "FFFFFF")
    c.fill      = hfill("1A237E")
    c.alignment = center()
    ws.row_dimensions[1].height = 38

    ws.append([])
    ws.merge_cells("A2:B2")
    c = ws["A2"]
    c.value     = candidate
    c.font      = Font(italic=True, size=11, color="37474F")
    c.fill      = hfill("E8EAF6")
    c.alignment = left()
    ws.row_dimensions[2].height = 30

    ws.append([])   # spacer

    ws.append(["Total Opportunities",  len(jobs)])
    ws.append(["🟢 High Probability",  len(high)])
    ws.append(["🟡 Medium Probability", len(medium)])
    ws.append(["🔴 Stretch Roles",      len(stretch)])
    ws.append(["📅 Target Date",        target])

    label_fills = ["283593", "1B5E20", "E65100", "B71C1C", "4A148C"]
    for i, fill_hex in enumerate(label_fills, 4):
        for col in [1, 2]:
            cell = ws.cell(row=i, column=col)
            cell.fill      = hfill(fill_hex)
            cell.font      = bfont(11, "FFFFFF")
            cell.alignment = left() if col == 2 else center()
            cell.border    = thin_border()
        ws.row_dimensions[i].height = 24

    ws.append([])

    ws.append(["✅ Resume Strengths", ""])
    c = ws.cell(row=ws.max_row, column=1)
    c.font = bfont(12, "1B5E20")
    c.fill = hfill("C8E6C9")
    c.border = thin_border()
    ws.merge_cells(f"A{ws.max_row}:B{ws.max_row}")
    for s in strengths:
        ws.append(["", f"• {s}"])
        ws.cell(row=ws.max_row, column=2).font = Font(size=11, color="1B1B1B")
        ws.row_dimensions[ws.max_row].height = 20

    ws.append([])

    ws.append(["⚠️ Resume Gaps", ""])
    c = ws.cell(row=ws.max_row, column=1)
    c.font = bfont(12, "B71C1C")
    c.fill = hfill("FFCDD2")
    c.border = thin_border()
    ws.merge_cells(f"A{ws.max_row}:B{ws.max_row}")
    for g in gaps:
        ws.append(["", f"• {g}"])
        ws.cell(row=ws.max_row, column=2).font = Font(size=11, color="1B1B1B")
        ws.row_dimensions[ws.max_row].height = 20

    # ── Sheet 2: All Jobs ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("🗂 All Jobs")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A3"

    ws2.append([])
    ws2.merge_cells(f"A1:{get_column_letter(len(JOB_HEADERS))}1")
    c = ws2["A1"]
    c.value = f"All Opportunities — Sorted by Fit Score  |  {today}"
    c.font = bfont(13)
    c.fill = hfill("1A237E")
    c.alignment = center()
    ws2.row_dimensions[1].height = 30

    header_row(ws2, JOB_HEADERS, "283593")
    for j in jobs:
        write_job_row(ws2, j)
    set_col_widths(ws2, JOB_WIDTHS)

    # ── Sheets 3–5: Categorised ───────────────────────────────────────────────
    for sheet_title, job_list, fill_title, fill_hdr in [
        ("🟢 High Probability",   high,    "1B5E20", "2E7D32"),
        ("🟡 Medium Probability", medium,  "E65100", "EF6C00"),
        ("🔴 Stretch Roles",      stretch, "B71C1C", "C62828"),
    ]:
        wsc = wb.create_sheet(sheet_title)
        wsc.sheet_view.showGridLines = False
        wsc.freeze_panes = "A3"

        wsc.append([])
        wsc.merge_cells(f"A1:{get_column_letter(len(JOB_HEADERS))}1")
        c = wsc["A1"]
        c.value = f"{sheet_title}  |  {len(job_list)} roles  |  {today}"
        c.font = bfont(13)
        c.fill = hfill(fill_title)
        c.alignment = center()
        wsc.row_dimensions[1].height = 30

        header_row(wsc, JOB_HEADERS, fill_hdr)
        for j in job_list:
            write_job_row(wsc, j)
        set_col_widths(wsc, JOB_WIDTHS)

    # ── Sheet 6: 30-Day Resume Tips ───────────────────────────────────────────
    ws6 = wb.create_sheet("📅 30-Day Resume Tips")
    ws6.sheet_view.showGridLines = False
    ws6.freeze_panes = "A3"

    ws6.append([])
    ws6.merge_cells("A1:F1")
    c = ws6["A1"]
    c.value = f"30-Day Resume Improvement Plan  |  {today} → {target}"
    c.font = bfont(13)
    c.fill = hfill("4A148C")
    c.alignment = center()
    ws6.row_dimensions[1].height = 30

    tip_headers = ["Day", "Date", "Focus Area", "Tip (Actionable)", "Before", "After"]
    header_row(ws6, tip_headers, "6A1B9A")

    for tip in tips:
        ws6.append([
            tip.get("day", ""),
            tip.get("date", ""),
            tip.get("focus_area", ""),
            tip.get("tip", ""),
            tip.get("example_before", ""),
            tip.get("example_after", ""),
        ])
        r = ws6.max_row
        ws6.row_dimensions[r].height = 44
        fill_hex = "F3E5F5" if tip.get("day", 0) % 2 == 0 else "FFFFFF"
        for col in range(1, 7):
            cell = ws6.cell(row=r, column=col)
            cell.fill      = hfill(fill_hex)
            cell.alignment = left()
            cell.border    = thin_border()
            cell.font      = Font(size=10, color="1B1B1B")
        ws6.cell(row=r, column=1).alignment = center()

    set_col_widths(ws6, [5, 12, 20, 52, 36, 36])

    wb.save(output_path)
    print(f"✅ Excel workbook saved → {output_path}")


# ── entrypoint ────────────────────────────────────────────────────────────────

def main():
    today       = datetime.date.today()
    target_date = today + datetime.timedelta(days=30)

    print("📄 Reading resume...")
    resume_text = read_resume()
    print(f"📝 Resume length: {len(resume_text)} chars")

    print("📝 Preparing prompt...")
    with open("prompt.txt", "r") as f:
        raw_prompt = f.read()

    # Strip comment lines
    prompt_lines = [l for l in raw_prompt.splitlines()
                    if not l.strip().startswith("#")]
    prompt = "\n".join(prompt_lines).strip()
    prompt = prompt.replace("{TODAY}",       str(today))
    prompt = prompt.replace("{TARGET_DATE}", str(target_date))
    prompt = prompt.replace("{RESUME_TEXT}", resume_text)

    print("🤖 Calling OpenRouter LLM...")
    data, model_used = call_openrouter(prompt, call_label="JOBS")
    print(f"✅ Response received from: {model_used}")

    output_path = f"jobs_{today}.xlsx"
    print("📊 Building Excel workbook...")
    build_excel(data, output_path)


if __name__ == "__main__":
    main()
